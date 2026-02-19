"""Export utilities for dashboard tables."""
from __future__ import annotations

from io import BytesIO
import re
from typing import Any, List, Optional, Tuple

import pandas as pd


class ExportBuilder:
    """Builds export files (Excel and PDF) from dashboard tables."""

    @staticmethod
    def build_excel_bytes(
        tables: List[Tuple[str, pd.DataFrame]],
        charts: Optional[List[Tuple[str, Any]]] = None,
    ) -> bytes:
        """Build an Excel file in a single sheet with vertical table/chart blocks."""
        from openpyxl.drawing.image import Image as XLImage

        output = BytesIO()
        chart_entries = list(charts or [])
        consumed_chart_indexes: set[int] = set()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            used_names = set()
            sheet_name = ExportBuilder._safe_sheet_name("Resumen KPI", used_names)
            used_names.add(sheet_name)
            current_start_row = 0

            for raw_name, table in tables:
                if table is None or table.empty:
                    continue

                export_table = table.reset_index().copy()
                section_title = pd.DataFrame([[raw_name]])
                section_title.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=current_start_row,
                    startcol=0,
                    index=False,
                    header=False,
                )

                table_start_row = current_start_row + 1
                export_table.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=table_start_row,
                    startcol=0,
                    index=False,
                )

                ws = writer.book[sheet_name]
                header_row = table_start_row + 1
                data_start_row = header_row + 1
                data_end_row = table_start_row + len(export_table) + 1
                start_col = 1
                end_col = export_table.shape[1]
                ExportBuilder._coerce_table_cell_types(
                    ws,
                    data_start_row=data_start_row,
                    data_end_row=data_end_row,
                    start_col=start_col,
                    end_col=end_col,
                )
                ExportBuilder._autofit_columns(
                    ws,
                    start_col=start_col,
                    end_col=end_col,
                    header_row=header_row,
                    data_start_row=data_start_row,
                    data_end_row=data_end_row,
                )

                chart_idx = ExportBuilder._pick_chart_index(raw_name, chart_entries, consumed_chart_indexes)
                if chart_idx is None:
                    current_start_row = data_end_row + 2
                    continue

                chart_name, chart_fig = chart_entries[chart_idx]
                consumed_chart_indexes.add(chart_idx)
                start_row = data_end_row + 2
                ws[f"A{start_row}"] = f"Gráfico: {chart_name}"

                image_bytes = ExportBuilder._figure_to_png_bytes(chart_fig)
                if image_bytes is not None:
                    img_stream = BytesIO(image_bytes)
                    image = XLImage(img_stream)
                    image.width = 1100
                    image.height = 400
                    ws.add_image(image, f"A{start_row + 1}")
                    current_start_row = start_row + 24
                    continue

                native_ok = ExportBuilder._add_native_excel_chart(ws, chart_fig, start_row + 1)
                if not native_ok:
                    ws[f"A{start_row + 1}"] = "No se pudo generar este gráfico en Excel."
                    current_start_row = start_row + 3
                    continue

                current_start_row = start_row + 20

        output.seek(0)
        return output.getvalue()

    @staticmethod
    def build_pdf_bytes(
        tables: List[Tuple[str, pd.DataFrame]],
        title: str,
        filters_text: str,
        charts: Optional[List[Tuple[str, Any]]] = None,
    ) -> bytes:
        """Build a PDF file with visible tables and optional charts."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Image as RLImage
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )
        styles = getSampleStyleSheet()
        chart_entries = list(charts or [])
        consumed_chart_indexes: set[int] = set()
        page_width = landscape(A4)[0] - (20 * mm)
        story = [
            Paragraph(title, styles["Title"]),
            Spacer(1, 4 * mm),
            Paragraph(filters_text, styles["Normal"]),
            Spacer(1, 5 * mm),
        ]

        for name, table in tables:
            if table is None or table.empty:
                continue
            story.append(Paragraph(name, styles["Heading3"]))

            export_table = table.reset_index().copy()
            export_table = export_table.fillna("").astype(str)
            header = export_table.columns.tolist()
            rows = export_table.values.tolist()
            matrix = [header] + rows

            col_widths = ExportBuilder._build_pdf_column_widths(header, rows, page_width)

            pdf_table = Table(matrix, repeatRows=1, colWidths=col_widths)
            pdf_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAEAEA")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            story.append(pdf_table)

            chart_idx = ExportBuilder._pick_chart_index(name, chart_entries, consumed_chart_indexes)
            if chart_idx is not None:
                chart_name, chart_fig = chart_entries[chart_idx]
                consumed_chart_indexes.add(chart_idx)
                story.append(Spacer(1, 2 * mm))
                story.append(Paragraph(f"Gráfico: {chart_name}", styles["Normal"]))
                chart_bytes = ExportBuilder._figure_to_pdf_image_bytes(chart_fig)
                if chart_bytes is None:
                    story.append(Paragraph("No se pudo renderizar este gráfico.", styles["Normal"]))
                else:
                    chart_image = RLImage(BytesIO(chart_bytes), width=240 * mm, height=82 * mm)
                    story.append(chart_image)

            story.append(Spacer(1, 6 * mm))

        doc.build(story)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _safe_sheet_name(raw_name: str, used_names: set[str]) -> str:
        """Generate a valid and unique Excel sheet name."""
        invalid = ["[", "]", "*", "?", "/", "\\", ":"]
        clean_name = raw_name
        for token in invalid:
            clean_name = clean_name.replace(token, "-")
        clean_name = clean_name.strip() or "Hoja"
        clean_name = clean_name[:31]

        if clean_name not in used_names:
            return clean_name

        suffix = 1
        while True:
            candidate = f"{clean_name[:28]}-{suffix}"
            if candidate not in used_names:
                return candidate
            suffix += 1

    @staticmethod
    def _figure_to_png_bytes(fig: Any) -> Optional[bytes]:
        """Convert a Plotly figure to PNG bytes for file export."""
        if fig is None:
            return None
        try:
            return fig.to_image(format="png", width=1200, height=420, scale=1)
        except Exception:
            return None

    @staticmethod
    def _figure_to_pdf_image_bytes(fig: Any) -> Optional[bytes]:
        """Convert a Plotly figure to compressed bytes optimized for PDF memory usage."""
        if fig is None:
            return None
        try:
            png_bytes = fig.to_image(format="png", width=900, height=320, scale=1)
        except Exception:
            return None

        try:
            from PIL import Image

            with Image.open(BytesIO(png_bytes)) as image:
                if image.mode in ("RGBA", "P"):
                    image = image.convert("RGB")
                optimized = BytesIO()
                image.save(optimized, format="JPEG", quality=70, optimize=True)
                optimized.seek(0)
                return optimized.getvalue()
        except Exception:
            return png_bytes

    @staticmethod
    def _pick_chart_index(
        table_name: str,
        charts: List[Tuple[str, Any]],
        consumed_indexes: set[int],
    ) -> Optional[int]:
        """Pick chart index using strict deterministic name matching."""
        normalized_table = ExportBuilder._normalized_name(table_name)

        for idx, (chart_name, chart_fig) in enumerate(charts):
            if idx in consumed_indexes or chart_fig is None:
                continue
            if ExportBuilder._normalized_name(chart_name) == normalized_table:
                return idx

        return None

    @staticmethod
    def _normalized_name(name: str) -> str:
        """Normalize a name for strict and deterministic matching."""
        normalized = str(name).lower().strip()
        replacements = {
            "á": "a",
            "é": "e",
            "í": "i",
            "ó": "o",
            "ú": "u",
        }
        for source, target in replacements.items():
            normalized = normalized.replace(source, target)
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized

    @staticmethod
    def _build_pdf_column_widths(
        header: List[Any],
        rows: List[List[Any]],
        total_width: float,
    ) -> List[float]:
        """Build column widths proportionally to content, constrained to page width."""
        col_count = max(len(header), 1)
        if col_count == 1:
            return [total_width]

        sample_rows = rows[:300]
        weights: List[float] = []
        for col_idx in range(col_count):
            header_len = len(str(header[col_idx])) if col_idx < len(header) else 0
            max_len = header_len
            for row in sample_rows:
                if col_idx >= len(row):
                    continue
                text = "" if row[col_idx] is None else str(row[col_idx]).replace("\n", " ")
                max_len = max(max_len, min(len(text), 60))
            weights.append(float(max(max_len, 6)))

        total_weight = sum(weights)
        if total_weight <= 0:
            return [total_width / col_count] * col_count

        min_width = 35.0
        max_width = 200.0
        scaled = [total_width * (weight / total_weight) for weight in weights]
        bounded = [min(max(width, min_width), max_width) for width in scaled]

        bounded_sum = sum(bounded)
        if bounded_sum <= 0:
            return [total_width / col_count] * col_count

        factor = total_width / bounded_sum
        return [width * factor for width in bounded]

    @staticmethod
    def _name_tokens(name: str) -> set[str]:
        """Extract normalized tokens from a table or chart name."""
        normalized = str(name).lower()
        replacements = {
            "á": "a",
            "é": "e",
            "í": "i",
            "ó": "o",
            "ú": "u",
        }
        for source, target in replacements.items():
            normalized = normalized.replace(source, target)

        parts = [part.strip() for part in normalized.replace("(", " ").replace(")", " ").split("-")]
        words = set()
        for part in parts:
            for token in part.split():
                if token:
                    words.add(token)
        return words

    @staticmethod
    def _add_native_excel_chart(ws: Any, fig: Any, start_row: int) -> bool:
        """Create an Excel-native line chart from Plotly data as fallback."""
        try:
            from openpyxl.chart import LineChart, Reference
            from openpyxl.utils import get_column_letter
        except Exception:
            return False

        traces = []
        for trace in getattr(fig, "data", []):
            x_raw = getattr(trace, "x", None)
            y_raw = getattr(trace, "y", None)
            x_values = list(x_raw) if x_raw is not None else []
            y_values = list(y_raw) if y_raw is not None else []
            if not x_values or not y_values:
                continue
            name = getattr(trace, "name", None) or f"Serie {len(traces) + 1}"
            traces.append((name, x_values, y_values))

        if not traces:
            return False

        source_start_col = 200
        source_end_col = source_start_col + len(traces)

        max_len = max(len(y_values) for _, _, y_values in traces)
        ws.cell(row=start_row, column=source_start_col, value="Periodo")
        for col_idx, (name, _, _) in enumerate(traces, start=source_start_col + 1):
            ws.cell(row=start_row, column=col_idx, value=str(name))

        base_x = traces[0][1]
        for row_offset in range(max_len):
            row_number = start_row + 1 + row_offset
            x_value = base_x[row_offset] if row_offset < len(base_x) else ""
            ws.cell(row=row_number, column=source_start_col, value=str(x_value))

            for col_idx, (_, _, y_values) in enumerate(traces, start=source_start_col + 1):
                value = y_values[row_offset] if row_offset < len(y_values) else None
                try:
                    numeric_value = float(value) if value is not None else None
                except (TypeError, ValueError):
                    numeric_value = None
                ws.cell(row=row_number, column=col_idx, value=numeric_value)

        for hidden_col in range(source_start_col, source_end_col + 1):
            ws.column_dimensions[get_column_letter(hidden_col)].hidden = True

        chart = LineChart()
        chart.title = "Tendencia"
        chart.height = 8
        chart.width = 22
        chart.plotVisOnly = False
        data_ref = Reference(
            ws,
            min_col=source_start_col + 1,
            min_row=start_row,
            max_col=source_end_col,
            max_row=start_row + max_len,
        )
        cats_ref = Reference(
            ws,
            min_col=source_start_col,
            min_row=start_row + 1,
            max_row=start_row + max_len,
        )
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{start_row}")
        return True

    @staticmethod
    def _coerce_table_cell_types(
        ws: Any,
        data_start_row: int,
        data_end_row: int,
        start_col: int,
        end_col: int,
    ) -> None:
        """Convert text-like numeric cells into Excel numeric types and formats."""
        for row in ws.iter_rows(
            min_row=data_start_row,
            max_row=data_end_row,
            min_col=start_col,
            max_col=end_col,
        ):
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                parsed = ExportBuilder._parse_numeric_text(cell.value)
                if parsed is None:
                    continue
                value, number_format = parsed
                cell.value = value
                if number_format:
                    cell.number_format = number_format

    @staticmethod
    def _parse_numeric_text(value: str) -> Optional[Tuple[float | int, str]]:
        """Parse string values to int/float/percent preserving expected formatting."""
        text = str(value).strip()
        if not text:
            return None

        percent_match = re.fullmatch(r"-?\d+(?:[\.,]\d+)?%", text)
        if percent_match:
            numeric_text = text[:-1].replace(".", "").replace(",", ".")
            try:
                percent_value = float(numeric_text) / 100
                return percent_value, "0.0%"
            except ValueError:
                return None

        if re.fullmatch(r"-?\d{1,3}(?:\.\d{3})+", text):
            try:
                return int(text.replace(".", "")), "#,##0"
            except ValueError:
                return None

        if re.fullmatch(r"-?\d+", text):
            try:
                return int(text), "#,##0"
            except ValueError:
                return None

        if re.fullmatch(r"-?\d+[\.,]\d+", text):
            normalized = text.replace(",", ".")
            try:
                decimals = len(normalized.split(".")[-1])
                number_format = "#,##0." + ("0" * min(decimals, 4))
                return float(normalized), number_format
            except ValueError:
                return None

        return None

    @staticmethod
    def _autofit_columns(
        ws: Any,
        start_col: int,
        end_col: int,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
    ) -> None:
        """Set column widths based on content length for header and table rows."""
        from openpyxl.utils import get_column_letter

        for col_idx in range(start_col, end_col + 1):
            max_length = 0
            header_value = ws.cell(row=header_row, column=col_idx).value
            if header_value is not None:
                max_length = max(max_length, len(str(header_value)))

            for row_idx in range(data_start_row, data_end_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is None:
                    continue
                if isinstance(cell_value, (int, float)):
                    text_len = len(f"{cell_value:,.2f}")
                else:
                    text_len = len(str(cell_value))
                max_length = max(max_length, text_len)

            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 40)
